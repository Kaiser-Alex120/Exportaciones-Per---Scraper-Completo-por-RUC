"""Genera el Excel final con los resultados del scraper.
Crea 3 hojas: Exportaciones, Resumen Empresas y Sin Página Web.
También puede importar un Excel descargado directamente desde Aduanet."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import re

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Paleta de colores para el Excel
AZUL_OSCURO   = "1F3864"
AZUL_MEDIO    = "2E75B6"
AZUL_CLARO    = "BDD7EE"
VERDE         = "70AD47"
ROJO          = "FF0000"
NARANJA       = "ED7D31"
AMARILLO_SUAVE= "FFF2CC"
GRIS_CLARO    = "F2F2F2"
BLANCO        = "FFFFFF"


def exportar_excel(resultados: list) -> str:
    """
    Exporta todos los resultados a un Excel con 3 hojas:
      1. Productos         → detalle completo
      2. Resumen Empresas  → estado de cada empresa
      3. Sin Página Web    → empresas sin web detectada

    Args:
        resultados: lista de dicts con estructura:
          {
            "ruc": str,
            "razon_social": str,
            "url": str | None,
            "estado_web": "ok" | "sin_productos" | "sin_web" | "error",
            "mensaje": str,
            "productos": [ {"nombre", "precio", "descripcion", "url_fuente"} ]
          }
    """
    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"PRODUCTOS_EMPRESAS_{fecha}.xlsx"
    ruta = os.path.join(OUTPUT_DIR, nombre_archivo)

    wb = Workbook()

    # Crear las 3 hojas en orden lógico: primero el resumen, luego el detalle
    _crear_hoja_resumen(wb, resultados)
    _crear_hoja_productos(wb, resultados)
    _crear_hoja_sin_web(wb, resultados)

    # Eliminar hoja por defecto vacía si existe
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(ruta)
    return ruta


# Hoja 1: Exportaciones
def _crear_hoja_productos(wb: Workbook, resultados: list):
    ws = wb.create_sheet("Exportaciones")

    # Título
    _escribir_titulo(ws, "EXPORTACIONES POR EMPRESA — PERÚ", 12)

    # Encabezados (eliminadas CIF, ADV, IMP. ARANCEL porque vienen vacías de Aduanet)
    encabezados = ["N°", "RUC", "RAZÓN SOCIAL", "WEB EMPRESA",
                   "DUA", "OPERADOR/IMPORTADOR", "MES", "AGENTE ADUANERO", "ADUANA", "PAÍS DESTINO",
                   "FOB (US$)"]
    fila_enc = 3
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=fila_enc, column=col, value=enc)
        _estilo_encabezado(celda, AZUL_OSCURO)

    # Datos
    fila = fila_enc + 1
    nro_global = 1
    thin = _borde_fino()

    for res in resultados:
        ruc         = res.get("ruc", "")
        razon       = res.get("razon_social", "")
        url         = res.get("url") or ""
        estado_web  = res.get("estado_web", "sin_web")
        productos   = res.get("productos", [])

        if not productos:
            # Empresa sin productos (sin web o sin detección)
            valores = [
                nro_global, ruc, razon, url,
                _etiqueta_estado(estado_web),
                "", "", "", "", "", ""
            ]
            color_fila = AMARILLO_SUAVE if estado_web == "sin_productos" else "FFD7D7"
            for col, val in enumerate(valores, 1):
                c = ws.cell(row=fila, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=color_fila)
                c.border = thin
                c.alignment = Alignment(wrap_text=True, vertical="center")
            fila += 1
            nro_global += 1
        else:
            # Ordenar productos por MES de más reciente a más antiguo
            productos_ordenados = _ordenar_por_mes(productos)
            for i, prod in enumerate(productos_ordenados):
                es_primera = (i == 0)
                valores = [
                    nro_global if es_primera else "",
                    ruc        if es_primera else "",
                    razon      if es_primera else "",
                    url        if es_primera else "",
                    prod.get("dua", ""),
                    prod.get("operador", ""),
                    prod.get("mes", ""),
                    prod.get("agente", ""),
                    prod.get("aduana", ""),
                    prod.get("pais", ""),
                    prod.get("fob", ""),
                ]
                color = GRIS_CLARO if nro_global % 2 == 0 else BLANCO
                for col, val in enumerate(valores, 1):
                    c = ws.cell(row=fila, column=col, value=val)
                    c.fill = PatternFill("solid", fgColor=color)
                    c.border = thin
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                fila += 1
            nro_global += 1

    # Ancho de columnas ajustado
    anchos = [6, 14, 38, 28, 12, 42, 12, 28, 14, 16, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = f"A{fila_enc + 1}"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[fila_enc].height = 20


# Hoja 2: Resumen de empresas
def _crear_hoja_resumen(wb: Workbook, resultados: list):
    ws = wb.create_sheet("Resumen Empresas")

    _escribir_titulo(ws, "RESUMEN ESTADO DE EMPRESAS", 11)

    encabezados = ["N°", "RUC", "RAZÓN SOCIAL", "LOCALIZACIÓN",
                   "URL ENCONTRADA", "TELÉFONO", "EMAIL", "REDES SOCIALES",
                   "ESTADO", "N° EXPORTACIONES", "OBSERVACIÓN"]
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=3, column=col, value=enc)
        _estilo_encabezado(celda, AZUL_MEDIO)

    thin = _borde_fino()
    for i, res in enumerate(resultados, 1):
        fila = i + 3
        estado = res.get("estado_web", "sin_web")
        n_prod = len(res.get("productos", []))
        pags   = len(res.get("paginas_visitadas", []))

        # Color según estado
        if estado == "ok":
            color = "E2EFDA"         # verde claro
            icono = "[OK] Web con exportaciones"
        elif estado == "sin_productos":
            color = "FFF2CC"         # amarillo
            icono = "[WARN] Web sin exportaciones"
        elif estado == "sin_web":
            color = "FFD7D7"         # rojo claro
            icono = "[X] Sin pagina web"
        else:
            color = GRIS_CLARO
            icono = "[ERR] Error de acceso"

        contactos = res.get("contactos", {})
        tel = ", ".join(contactos.get("telefonos", [])[:2]) or "—"
        mail = ", ".join(contactos.get("emails", [])[:2]) or "—"
        direc = (contactos.get("direccion", "") or "")[:70] or "—"
        rs = ", ".join(contactos.get("redes_sociales", [])[:2]) or "—"

        valores = [
            i,
            res.get("ruc", ""),
            res.get("razon_social", ""),
            res.get("direccion_sunat", "—") or "—",
            res.get("url") or "—",
            tel,
            mail,
            rs,
            icono,
            n_prod if n_prod else "—",
            res.get("mensaje", ""),
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=color)
            c.border = thin
            c.alignment = Alignment(wrap_text=True, vertical="center")

    anchos = [5, 14, 38, 38, 30, 18, 30, 35, 22, 16, 40]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "A4"


# Hoja 3: Empresas sin página web
def _crear_hoja_sin_web(wb: Workbook, resultados: list):
    ws = wb.create_sheet("Sin Página Web")

    sin_web = [r for r in resultados if r.get("estado_web") == "sin_web"]

    _escribir_titulo(ws, f"EMPRESAS SIN PÁGINA WEB DETECTADA  ({len(sin_web)} empresas)", 4)

    encabezados = ["N°", "RUC", "RAZÓN SOCIAL", "OBSERVACIÓN"]
    for col, enc in enumerate(encabezados, 1):
        celda = ws.cell(row=3, column=col, value=enc)
        _estilo_encabezado(celda, "C00000")   # rojo oscuro

    thin = _borde_fino()
    for i, res in enumerate(sin_web, 1):
        fila = i + 3
        color = GRIS_CLARO if i % 2 == 0 else "FFD7D7"
        valores = [
            i,
            res.get("ruc", ""),
            res.get("razon_social", ""),
            "No cuenta con página web",
        ]
        for col, val in enumerate(valores, 1):
            c = ws.cell(row=fila, column=col, value=val)
            c.fill = PatternFill("solid", fgColor=color)
            c.border = thin
            c.alignment = Alignment(wrap_text=True, vertical="center")

    anchos = [5, 15, 55, 35]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "A4"


# Utilidades de estilo para el Excel
def _escribir_titulo(ws, texto: str, n_cols: int):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    celda = ws.cell(row=1, column=1, value=texto)
    celda.font = Font(bold=True, size=13, color=BLANCO)
    celda.fill = PatternFill("solid", fgColor=AZUL_OSCURO)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    fecha_celda = ws.cell(row=2, column=1,
                          value=f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Fuente: SUNAT + Web empresas")
    fecha_celda.font = Font(italic=True, size=9, color=AZUL_OSCURO)
    fecha_celda.fill = PatternFill("solid", fgColor=AZUL_CLARO)
    fecha_celda.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16


def _estilo_encabezado(celda, color_fondo: str):
    celda.fill = PatternFill("solid", fgColor=color_fondo)
    celda.font = Font(bold=True, color=BLANCO, size=10)
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    celda.border = _borde_fino()


def _borde_fino() -> Border:
    lado = Side(style="thin", color="AAAAAA")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _etiqueta_estado(estado: str) -> str:
    mapa = {
        "ok":            "[OK] Web con exportaciones",
        "sin_productos": "[WARN] Web sin exportaciones",
        "sin_web":       "[X] Sin pagina web",
        "error":         "[ERR] Error al acceder",
    }
    return mapa.get(estado, estado)


def _simplificar_url_aduanet(url: str) -> str:
    """Simplifica la URL de Aduanet mostrando solo el dominio base y parámetros clave."""
    if not url or "aduanet" not in url:
        return url
    # Extraer solo la parte base sin todos los parámetros de paginación
    if "?" in url:
        base = url.split("?")[0]
        return f"{base} (Ver en Aduanet)"
    return url


def _ordenar_por_mes(productos: list) -> list:
    """Ordena productos por MES de más reciente a más antiguo.
    
    Formato esperado: 'Abril 2024', 'Septiembre 2025', etc.
    """
    # Mapa de meses español a número
    meses_es = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
        'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    
    def _extraer_fecha_key(prod):
        mes_str = prod.get('mes', '')
        if not mes_str:
            return (0, 0)
        
        # Parsear formato "Mes Año" (ej: "Abril 2024")
        partes = mes_str.lower().split()
        if len(partes) >= 2:
            mes_nombre = partes[0]
            try:
                anio = int(partes[1])
                mes_num = meses_es.get(mes_nombre, 0)
                return (anio, mes_num)
            except (ValueError, IndexError):
                pass
        return (0, 0)
    
    # Ordenar de más reciente a más antiguo (descendente)
    return sorted(productos, key=_extraer_fecha_key, reverse=True)


# Importar Excel descargado desde el botón "Exportar XLS" de Aduanet
def importar_excel_aduanet(ruta_excel_aduanet: str, datos_empresa: dict) -> str:
    """
    Importa Excel descargado desde el boton "Exportar XLS" de Aduanet
    y lo combina con datos SUNAT + Web de la empresa.
    
    Args:
        ruta_excel_aduanet: Ruta al archivo Excel descargado de Aduanet
        datos_empresa: Dict con datos de SUNAT y Web
            {
                "ruc": "...",
                "razon_social": "...",
                "url": "...",
                "contactos": {"telefonos": [], "emails": [], "redes_sociales": []},
                "direccion_sunat": "..."
            }
    
    Returns:
        Dict con la ruta al Excel combinado y la cantidad de registros importados.
        {
          "ruta": "...",
          "n_productos": 0
        }
    """
    print(f"[INFO] Importando Excel de Aduanet: {ruta_excel_aduanet}")
    
    # Leer Excel de Aduanet
    try:
        df_aduanet = pd.read_excel(ruta_excel_aduanet)
        print(f"[OK] Leidos {len(df_aduanet)} registros de Aduanet")
    except Exception as e:
        print(f"[ERROR] No se pudo leer Excel de Aduanet: {e}")
        return None
    
    # Preparar datos de la empresa para el formato de resultados
    resultado = {
        "ruc": datos_empresa.get("ruc", ""),
        "razon_social": datos_empresa.get("razon_social", ""),
        "url": datos_empresa.get("url", ""),
        "direccion_sunat": datos_empresa.get("direccion_sunat", ""),
        "estado_web": "ok" if datos_empresa.get("url") else "sin_web",
        "mensaje": f"{len(df_aduanet)} registros importados de Aduanet",
        "contactos": datos_empresa.get("contactos", {
            "telefonos": [],
            "emails": [],
            "direccion": "",
            "redes_sociales": []
        }),
        # Convertir DataFrame de Aduanet a lista de productos
        "productos": _df_aduanet_a_productos(df_aduanet),
        "paginas_visitadas": [datos_empresa.get("url", "")] if datos_empresa.get("url") else [],
        "anios": [datetime.now().year]
    }
    
    # Generar Excel combinado
    ruta = exportar_excel([resultado])
    return {
        "ruta": ruta,
        "n_productos": len(resultado["productos"])
    }


def _df_aduanet_a_productos(df: pd.DataFrame) -> list:
    """Convierte DataFrame de Aduanet a lista de productos estandarizada."""
    productos = []
    
    # Mapeo de columnas típicas de Aduanet
    columnas = df.columns.tolist()
    print(f"[INFO] Columnas detectadas: {columnas}")
    
    for _, row in df.iterrows():
        prod = {
            "dua": str(row.get("DUA", "")) if "DUA" in columnas else "",
            "operador": str(row.get("IMPORTADOR", row.get("OPERADOR", ""))),
            "mes": str(row.get("MES", "")) if "MES" in columnas else "",
            "agente": str(row.get("AGENTE", "")) if "AGENTE" in columnas else "",
            "aduana": str(row.get("ADUANA", "")) if "ADUANA" in columnas else "",
            "pais": str(row.get("PAIS", row.get("PAÍS", ""))),
            "fob": str(row.get("FOB", "")) if "FOB" in columnas else "",
            "cif": str(row.get("CIF", "")) if "CIF" in columnas else "",
            "adv": str(row.get("ADV", "")) if "ADV" in columnas else "",
            "imp_arancel": str(row.get("IMP. ARANCEL", row.get("IMP_ARANCEL", ""))),
            "url_fuente": "http://www.aduanet.gob.pe (Importado de Excel)",
            "anio": datetime.now().year
        }
        productos.append(prod)
    
    return productos